import openpyxl

wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
print(f"Sheets: {wb.sheetnames}")

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\n--- Sheet: {name} ({sheet.max_row} rows, {sheet.max_column} cols) ---")
    
    # Print the first 20 rows and 15 columns
    for r in range(1, min(25, sheet.max_row + 1)):
        row_vals = []
        for c in range(1, min(15, sheet.max_column + 1)):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if val is None:
                val = ""
            row_vals.append(str(val))
        print(f"Row {r:2d}: {row_vals}")
