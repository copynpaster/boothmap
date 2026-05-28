import re

# Read data.js to parse RAW_EXHIBITORS
raw_exhibitors_ids = set()
with open("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/data.js", "r", encoding="utf-8") as f:
    content = f.read()
    # Find all id: "..." in RAW_EXHIBITORS
    matches = re.findall(r'id:\s*["\']([^"\']+)["\']', content)
    for m in matches:
        if m != "무대":
            raw_exhibitors_ids.add(m)

# Read Excel sheet
import openpyxl
wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

excel_ids = set()
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if val:
            val_str = str(val).strip()
            # Clean and match A-H booth IDs
            val_clean = val_str.replace("I", "1").replace("Z", "7").replace("O", "0").replace("S", "5")
            if re.match(r"^[A-H]\d+$", val_clean):
                excel_ids.add(val_clean)

print(f"Total IDs in RAW_EXHIBITORS: {len(raw_exhibitors_ids)}")
print(f"Total IDs in Excel: {len(excel_ids)}")

missing_in_excel = raw_exhibitors_ids - excel_ids
missing_in_raw = excel_ids - raw_exhibitors_ids

print(f"\nMissing in Excel (in RAW_EXHIBITORS but not in Excel): {sorted(list(missing_in_excel))}")
print(f"Missing in RAW_EXHIBITORS (in Excel but not in RAW_EXHIBITORS): {sorted(list(missing_in_raw))}")
