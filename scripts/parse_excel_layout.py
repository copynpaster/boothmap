import openpyxl
import json

wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

max_r = sheet.max_row
max_c = sheet.max_column

# Parse merged cell ranges
merged_lookup = {}  # maps (r, c) to (r_start, r_end, c_start, c_end)
for merged_range in sheet.merged_cells.ranges:
    r_start, c_start, r_end, c_end = merged_range.bounds[1], merged_range.bounds[0], merged_range.bounds[3], merged_range.bounds[2]
    for r in range(r_start, r_end + 1):
        for c in range(c_start, c_end + 1):
            merged_lookup[(r, c)] = (r_start, r_end, c_start, c_end)

booth_positions = {}

for r in range(1, max_r + 1):
    for c in range(1, max_c + 1):
        val = sheet.cell(row=r, column=c).value
        if not val:
            continue
        
        # If it's a merged cell, get its dimensions
        if (r, c) in merged_lookup:
            r_start, r_end, c_start, c_end = merged_lookup[(r, c)]
            # Only process if we are at the top-left of the merged range
            if r == r_start and c == c_start:
                gridX = c_start - 1
                gridW = c_end - c_start + 1
                gridY = 27 - r_end
                gridH = r_end - r_start + 1
            else:
                continue
        else:
            gridX = c - 1
            gridW = 1
            gridY = 27 - r
            gridH = 1
        
        # Save booth position
        booth_id = str(val).strip()
        if booth_id:
            booth_positions[booth_id] = {
                'gridX': gridX,
                'gridY': gridY,
                'gridW': gridW,
                'gridH': gridH
            }

# Let's print the generated positions sorted by booth ID
print("Generated Booth Grid Mapping:")
print(json.dumps(booth_positions, indent=2, ensure_ascii=False))

# Write to a JSON file so we can easily read it in data.js
with open("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/excel_grid_mapping.json", "w", encoding="utf-8") as f:
    json.dump(booth_positions, f, indent=2, ensure_ascii=False)
