import openpyxl

wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

pillars = []
entrances = []

for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.cell(row=r, column=c).value
        if val:
            val_str = str(val).strip()
            if val_str == "기둥":
                pillars.append({
                    'gridX': c - 1,
                    'gridY': 27 - r
                })
            elif "입구" in val_str or "출구" in val_str or "등록대" in val_str:
                entrances.append({
                    'name': val_str,
                    'gridX': c - 1,
                    'gridY': 27 - r
                })

print("Pillars found:")
print(pillars)
print("\nEntrances found:")
print(entrances)
