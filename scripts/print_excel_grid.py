import openpyxl

wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

# Let's inspect the entire sheet size
max_r = sheet.max_row
max_c = sheet.max_column
print(f"Excel Grid Dimensions: {max_r} rows x {max_c} cols")

# Print the grid in a clean layout
for r in range(1, max_r + 1):
    row_cells = []
    for c in range(1, max_c + 1):
        val = sheet.cell(row=r, column=c).value
        if val is None:
            val = "."
        row_cells.append(f"{str(val):8s}")
    print(f"Row {r:02d}: " + " ".join(row_cells))
