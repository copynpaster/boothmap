import re
import openpyxl
import json

# 1. Read data.js and extract RAW_EXHIBITORS
with open("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/data.js", "r", encoding="utf-8") as f:
    content = f.read()

# Extract from line containing "const RAW_EXHIBITORS = [" to the closing "];"
# We can use a regex to capture it
raw_exhibitors_match = re.search(r"const RAW_EXHIBITORS = \[(.*?)\n\];", content, re.DOTALL)
if not raw_exhibitors_match:
    print("Error: Could not find RAW_EXHIBITORS in data.js")
    exit(1)

raw_exhibitors_text = raw_exhibitors_match.group(1)

# 2. Parse Excel sheet
wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

max_r = sheet.max_row
max_c = sheet.max_column

# Parse merged cell ranges
merged_lookup = {}
for merged_range in sheet.merged_cells.ranges:
    r_start, c_start, r_end, c_end = merged_range.bounds[1], merged_range.bounds[0], merged_range.bounds[3], merged_range.bounds[2]
    for r in range(r_start, r_end + 1):
        for c in range(c_start, c_end + 1):
            merged_lookup[(r, c)] = (r_start, r_end, c_start, c_end)

booth_positions = {}
pillars = []
special_labels = [] # for 입구, 출구, 등록대

for r in range(1, max_r + 1):
    for c in range(1, max_c + 1):
        val = sheet.cell(row=r, column=c).value
        if not val:
            continue
        
        val_str = str(val).strip()
        
        # Merge info
        if (r, c) in merged_lookup:
            r_start, r_end, c_start, c_end = merged_lookup[(r, c)]
            if r == r_start and c == c_start:
                gridX = c_start - 1
                gridW = c_end - c_start + 1
                gridY = 27 - r_end
                gridH = r_end - r_start + 1
            else:
                continue
        else:
            gridX = c - 1
            gridW = 1
            gridY = 27 - r
            gridH = 1
        
        # Clean booth ID
        val_clean = val_str.replace("I", "1").replace("Z", "7").replace("O", "0").replace("S", "5")
        
        if val_str == "기둥":
            pillars.append({
                'gridX': gridX,
                'gridY': gridY,
                'gridW': gridW,
                'gridH': gridH
            })
        elif val_str in ["입구", "출구", "등록대"]:
            # We will assign unique IDs to exits and entrances to render them as static cells
            # E.g. 입구_1, 입구_2
            special_labels.append({
                'name': val_str,
                'gridX': gridX,
                'gridY': gridY,
                'gridW': gridW,
                'gridH': gridH
            })
        else:
            # Regular booth or 무대
            booth_positions[val_clean] = {
                'gridX': gridX,
                'gridY': gridY,
                'gridW': gridW,
                'gridH': gridH
            }

# Add the 5 missing booths with their calculated positions
# G8 -> Row 18, Col 26 (gridX=25, gridY=9)
# G28 -> Row 3, Col 26 (gridX=25, gridY=24)
# B29 -> Row 1, Col 10 (gridX=9, gridY=26)
# C7 -> Row 19, Col 13 (gridX=12, gridY=8)
# D28 -> Row 6, Col 16 (gridX=15, gridY=21)
missing_mappings = {
    "G8": {"gridX": 25, "gridY": 9, "gridW": 1, "gridH": 1},
    "G28": {"gridX": 25, "gridY": 24, "gridW": 1, "gridH": 1},
    "B29": {"gridX": 9, "gridY": 26, "gridW": 1, "gridH": 1},
    "C7": {"gridX": 12, "gridY": 8, "gridW": 1, "gridH": 1},
    "D28": {"gridX": 15, "gridY": 21, "gridW": 1, "gridH": 1}
}

for bid, pos in missing_mappings.items():
    booth_positions[bid] = pos

# Let's add the special label booths as well so they render on the map!
# To render them cleanly without showing up as "미배정 부스" in search list,
# we can prefix their IDs (e.g. "ENTRANCE1", "EXIT1", etc.) and handle them.
entrance_count = 1
exit_count = 1
for label in special_labels:
    name = label['name']
    if name == "입구":
        bid = f"입구_{entrance_count}"
        entrance_count += 1
    elif name == "출구":
        bid = f"출구_{exit_count}"
        exit_count += 1
    else:
        bid = "등록대"
    
    booth_positions[bid] = {
        'gridX': label['gridX'],
        'gridY': label['gridY'],
        'gridW': label['gridW'],
        'gridH': label['gridH']
    }

# 3. Generate JS File Content
js_content = f"""// Raw Exhibitor List Transcribed from http://www.teanews.com/attach/20260526/1779764692.jpg
const RAW_EXHIBITORS = [{raw_exhibitors_text}
];

// Excel-based Grid Layout Mapping
const BOOTH_GRID_MAP = {json.dumps(booth_positions, indent=2, ensure_ascii=False)};

// Helper function to map booth ID to grid coordinates and cell dimensions
function getBoothGrid(id) {{
  return BOOTH_GRID_MAP[id] || null;
}}

// Helper function to convert grid units to pixels
function getBoothCoordinatesForGrid(gridX, gridY, gridW, gridH) {{
  const COL_UNIT = 28; // cell width (26px) + gap (2px)
  const ROW_UNIT = 22; // cell height (20px) + gap (2px)
  const margin_left = 32;
  const margin_top = 23;
  
  const x = margin_left + gridX * COL_UNIT;
  // gridY = 26 is Row 1 (top of SVG), gridY = 0 is Row 27 (bottom of SVG)
  const y = margin_top + (26 - (gridY + gridH - 1)) * ROW_UNIT;
  
  const width = gridW * 26 + (gridW - 1) * 2;
  const height = gridH * 20 + (gridH - 1) * 2;
  
  return {{ x, y, width, height, gridX, gridY, gridW, gridH }};
}}

// Generate coordinates dynamically based on booth ID
function getBoothCoordinates(id) {{
  const grid = getBoothGrid(id);
  if (!grid) return null;
  return getBoothCoordinatesForGrid(grid.gridX, grid.gridY, grid.gridW, grid.gridH);
}}

// Compile final list with coordinates
const ALL_BOOTHS = [];

// Helper list of all sections and their booth limits
const SECTIONS = {{
  'A': 49, 'B': 29, 'C': 31, 'D': 34, 'E': 34, 'F': 32, 'G': 28, 'H': 27
}};

const exhibitorsMap = {{}};
RAW_EXHIBITORS.forEach(ex => {{
  exhibitorsMap[ex.id] = ex.name;
}});

// Compile booths including empty ones if they exist in the grid map
for (const [id, grid] of Object.entries(BOOTH_GRID_MAP)) {{
  if (id === "무대" || id.startsWith("입구") || id.startsWith("출구") || id === "등록대") {{
    continue;
  }}
  const name = exhibitorsMap[id] || "미배정 부스";
  ALL_BOOTHS.push({{ id, name }});
}}

const DEFAULT_EXHIBITORS = ALL_BOOTHS.map(booth => {{
  const coords = getBoothCoordinates(booth.id);
  return {{
    id: booth.id,
    name: booth.name,
    section: booth.id.charAt(0),
    ...coords
  }};
}});

// Add Special Entities
DEFAULT_EXHIBITORS.push(
  {{ id: "무대", name: "메인 무대 (차 시연 및 공연)", section: "SPECIAL", ...getBoothCoordinates("무대") }}
);

// Add entrances, exits, registration desks as special static entities
const SPECIAL_ENTITIES = [
  {{ id: "입구_1", name: "입구", section: "SPECIAL" }},
  {{ id: "입구_2", name: "입구", section: "SPECIAL" }},
  {{ id: "출구_1", name: "출구", section: "SPECIAL" }},
  {{ id: "출구_2", name: "출구", section: "SPECIAL" }},
  {{ id: "등록대", name: "등록대", section: "SPECIAL" }}
];

SPECIAL_ENTITIES.forEach(entity => {{
  const coords = getBoothCoordinates(entity.id);
  if (coords) {{
    DEFAULT_EXHIBITORS.push({{ ...entity, ...coords }});
  }}
}});

// Pillar Coordinates
const PILLARS = {json.dumps(pillars, indent=2)}.map(p => {{
  const coords = getBoothCoordinatesForGrid(p.gridX, p.gridY, p.gridW, p.gridH);
  return {{ ...p, ...coords }};
}});
"""

with open("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/data.js", "w", encoding="utf-8") as f:
    f.write(js_content)

print("data.js successfully written!")
