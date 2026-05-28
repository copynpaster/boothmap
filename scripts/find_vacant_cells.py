import openpyxl

wb = openpyxl.load_workbook("/Users/uchun1ee/.gemini/antigravity/scratch/teaworld2026/booth_layout.xlsx")
sheet = wb.active

sections_cols = {
    'B': [9, 10],
    'C': [12, 13],
    'D': [15, 16],
    'E': [18, 19],
    'F': [21, 22],
    'G': [24, 25],
    'H': [28, 29]
}

# Find all vacant cells in these column ranges
# Exclude rows that are corridors (Rows 2, 5, 11, 17, 20, 26 in Excel? Let's check which rows have all empty cells)
corridor_rows = []
for r in range(1, sheet.max_row + 1):
    non_empty = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1) if sheet.cell(row=r, column=c).value]
    if not non_empty:
        corridor_rows.append(r)
print(f"Corridor rows: {corridor_rows}")

vacant_cells = []
for sec, cols in sections_cols.items():
    for r in range(1, sheet.max_row + 1):
        if r in corridor_rows:
            continue
        for c in cols:
            val = sheet.cell(row=r, column=c).value
            if val is None:
                # Check if it is covered by a merged range
                is_merged = False
                for merged_range in sheet.merged_cells.ranges:
                    r_start, c_start, r_end, c_end = merged_range.bounds[1], merged_range.bounds[0], merged_range.bounds[3], merged_range.bounds[2]
                    if r_start <= r <= r_end and c_start <= c <= c_end:
                        is_merged = True
                        break
                if not is_merged:
                    vacant_cells.append((sec, r, c))

print("\nVacant cells found:")
for sec, r, c in vacant_cells:
    col_name = openpyxl.utils.get_column_letter(c)
    print(f"Section {sec}: Excel Cell {col_name}{r} (Row {r}, Col {c})")
